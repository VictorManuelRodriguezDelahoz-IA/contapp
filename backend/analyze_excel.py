import openpyxl

# Load the Excel file
wb = openpyxl.load_workbook('../Finanzas Personales - 2026 (1).xlsx')

print("Sheets:", wb.sheetnames)
print("\n" + "="*80)

# Analyze ENERO sheet
ws = wb['ENERO']

print("\nFirst 20 rows of ENERO sheet:")
print("="*80)

for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    if i <= 20:
        print(f"Row {i}: {row}")

print("\n" + "="*80)
print(f"\nTotal rows in ENERO: {ws.max_row}")
print(f"Total columns in ENERO: {ws.max_column}")
